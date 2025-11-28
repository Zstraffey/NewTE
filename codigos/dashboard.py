import random
from datetime import datetime

import bcrypt
from PyQt5.QtCore import QTimer, Qt, QThread, pyqtSignal, QByteArray, QBuffer, QIODevice, QSize, QRect, QRectF, QDate
from PyQt5.QtWidgets import QMainWindow, QWidget, QPushButton, QMessageBox, QFileDialog, QTableWidgetItem, QHeaderView, \
    QSizePolicy, QGridLayout, QDialog, QVBoxLayout
from PyQt5.uic import loadUi
from PyQt5.QtGui import QPixmap, QIcon, QPainterPath, QRegion, QTextCharFormat, QBrush, QColor
from flask import session
import os

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

from openpyxl import Workbook

import calendar

class GraficoCanvas(FigureCanvas):
    def __init__(self, parent=None):
        self.fig = Figure(figsize=(5, 4), tight_layout=True)
        self.ax = self.fig.add_subplot(111)

        super().__init__(self.fig)
        self.setParent(parent)

        # Permite resize automático dentro do widget
        self.fig.set_facecolor("#FFFFFF")

import imgs_qrc
import mysql.connector as mc
import time
from functools import partial
from login import quitProgram

from codigos.classes import Session, bancoDados, ChatBubble, PopupSobreMim, PopupCargo, PopupDepto, PopupCalendario, ValidadorCPF, ValidadorRG, ValidadorEmail, GeradorSenha, PopupVisualizarCal, resource_path

import re
import unicodedata
import difflib

class EmailSender(QThread):
    finished = pyqtSignal(str)

    def __init__(self, receiver_email, senha):
        super().__init__()
        self.receiver_email = receiver_email
        self.senha = senha

    def run(self):
        try:
            sender_email = "newtetcc2025@gmail.com"
            app_password = "bboq pkqm nexm riyv"

            message = MIMEMultipart("related")
            message["From"] = sender_email
            message["To"] = self.receiver_email
            message["Subject"] = "New TE - Cadastro de Usuário"

            html_content = f"""
            <html>
            <head>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                    }}
                    .header {{
                        background-color: #6a1b9a;  /* Cor de fundo roxa */
                        color: white;
                        padding: 20px;
                        text-align: center;
                        font-size: 24px;
                        font-weight: bold;
                    }}
                    .sub-header {{
                        font-size: 16px;
                        color: #333;
                        text-align: center;
                        margin-top: 10px;
                    }}
                    .code {{
                        font-size: 20px;
                        font-weight: bold;
                        color: #333;
                        background-color: #f0f0f0;
                        padding: 10px;
                        border-radius: 5px;
                    }}
                    .centered-img {{
                        display: block;
                        margin-left: auto;
                        margin-right: auto;
                        width: 5%; /* Ajuste a largura da imagem conforme necessário */
                    }}
                </style>
            </head>
            <body>
                <!-- Cabeçalho -->

               <img src="cid:logo_image" alt="Logo da Empresa" class="centered-img" />

                <div class="header">
                    New TE - Onboarding Digital
                </div>

                <!-- Subcabeçalho -->
                <div class="sub-header">
                    Cadastro de Usuário
                </div>

                <p>Este e-mail é automaticamente gerado, não responda. A sua conta foi cadastrada com sucesso!</p>
                <p>O seu e-mail de entrada é: <span class="code">{str(self.receiver_email)}</span></p>
                <p>A sua senha automaticamente gerada é: <span class="code">{str(self.senha)}</span></p>
                <p>Caso queira mudar sua senha, entre com uma requisição de Recuperação de Senha.</p>
            </body>
            </html>
            """

            message.attach(MIMEText(html_content, "html"))
            image_path = resource_path("imagens/logo_new_te.png")
            if os.path.exists(image_path):
                with open(image_path, 'rb') as img_file:
                    img = MIMEImage(img_file.read())
                    img.add_header('Content-ID', '<logo_image>')
                    img.add_header('Content-Disposition', 'inline', filename="image.jpg")
                    message.attach(img)

            # Connect and send
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, app_password)
                server.sendmail(sender_email, self.receiver_email, message.as_string())

        except Exception as e:
            self.finished.emit(f"Erro ao enviar: {str(e)}")
            print(f"Erro ao enviar: {str(e)}")

leet_map = {
    'a': ['4', '@'],
    'e': ['3'],
    'i': ['1', '!', '|', 'l'],
    'o': ['0'],
    'u': ['v'],
    's': ['5', '$'],
    'g': ['9'],
    'b': ['8'],
    't': ['7'],
    'c': ['('],
}

# Lista de palavras ofensivas
palavras_bloqueadas = [
    "nigger", "nigga", "merda", "cu", "cus", "cuzao", "cuzão", "cv", "pcc",
    "foda", "fodo", "fodao", "fodão", "bundao", "bundão", "bunda", "bundinha",
    "viado", "bicha", "traveco", "tranny", "puto", "puta", "fdp", "filho da puta",
    "filha da puta", "urtiga", "desgraçado", "desgracado", "vai tomar no cu",
    "retardado", "mongol", "imbecil", "caralho", "krl", "putinho", "putinha",
    "nazista", "nazi", "arrombado", "arrombada", "crl"
]

def normalizar(texto):
    # Remove acentos e coloca tudo em minúsculo
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')
    return texto.lower()

def gerar_variacoes(palavra):
    # Cria regex com variações leet
    regex = ""
    for letra in palavra:
        letras_possiveis = [letra]
        if letra in leet_map:
            letras_possiveis += leet_map[letra]
        grupo = "[" + re.escape("".join(set(letras_possiveis))) + "]"
        regex += grupo
    return regex

def palavras_semelhantes(palavra, lista, limiar=0.8):
    # Retorna palavras da lista que são parecidas com a fornecida (acima de certo limiar)
    semelhantes = []
    for proibida in lista:
        ratio = difflib.SequenceMatcher(None, palavra, proibida).ratio()
        if ratio >= limiar:
            semelhantes.append(proibida)
    return semelhantes

def filtrar_texto(texto_original):
    texto_filtrado = texto_original
    texto_normalizado = normalizar(texto_original)

    # Divide o texto normalizado em palavras individuais para verificação de similaridade
    palavras_no_texto = re.findall(r'\b\w+\b', texto_normalizado)

    for palavra_texto in palavras_no_texto:
        similares = palavras_semelhantes(palavra_texto, palavras_bloqueadas)

        for proibida in similares:
            padrao_regex = gerar_variacoes(proibida)
            regex_compilado = re.compile(padrao_regex, re.IGNORECASE)

            # Substitui variações encontradas
            texto_filtrado = regex_compilado.sub(lambda m: m.group(0)[0] + "#" * (len(m.group(0)) - 1), texto_filtrado)

            # Também substitui se a palavra for parecida diretamente (ex: "desagraçado")
            texto_filtrado = re.sub(rf'\b{re.escape(palavra_texto)}\b',
                                    palavra_texto[0] + "#" * (len(palavra_texto) - 1),
                                    texto_filtrado, flags=re.IGNORECASE)

    if texto_filtrado != texto_original:
        QMessageBox.warning(None, "Aviso", "O termo utilizado na sua mensagem não se enquadra na Política de Linguagem e Conduta no ambiente de trabalho.")

    return texto_filtrado

class usuarioChat(QWidget):
    def __init__(self, user):#, callback):
        super().__init__()
        loadUi(resource_path("design/templates/contatos.ui"), self)

        self.nome_salvo.setText(user["nome"])

class licao(QWidget):
    def __init__(self):#, callback):
        super().__init__()
        loadUi(resource_path("design/templates/template_aula.ui"), self)

class adicionarLicao(QWidget):
    def __init__(self):#, callback):
        super().__init__()
        loadUi(resource_path("design/templates/aula_adicionar.ui"), self)

class TelaInicial(QMainWindow):
    class DBLoopUpdate(QThread):
        new_data = pyqtSignal(list)

        def __init__(self):
            super().__init__()
            self.running = True

        def query(self):
            if Session.current_user is None:
                self.stop()
                return

            db = bancoDados.conectar()
            if not db:
                return
            cursor = db.cursor()

            query = """
                 SELECT *
                 FROM mensagens_chat
                 WHERE id_mensagem > %s AND (
                     (remetente_id = %s AND destinatario_id = %s)
                     OR
                     (remetente_id = %s AND destinatario_id = %s)
                 )
                 ORDER BY data_envio ASC;
             """
            cursor.execute(query, (Session.last_message_id, Session.current_user["id_user"], Session.loaded_chat, Session.loaded_chat, Session.current_user["id_user"]))
            results = cursor.fetchall()
            if results:
                query = """
                     UPDATE mensagens_chat
                     SET lida = 1
                     WHERE lida = 0 AND destinatario_id = %s AND remetente_id = %s
                 """
                cursor.execute(query, (Session.current_user["id_user"], Session.loaded_chat))
                db.commit()
                print(query)

                self.new_data.emit(results)

        def run(self):
            while self.running:
                if Session.current_user is None:
                    self.stop()
                    return

                db = bancoDados.conectar()
                if not db:
                    return
                cursor = db.cursor()

                query = """
                     SELECT *
                     FROM mensagens_chat
                     WHERE id_mensagem > %s AND (
                         (remetente_id = %s AND destinatario_id = %s)
                         OR
                         (remetente_id = %s AND destinatario_id = %s)
                     )
                     ORDER BY data_envio ASC;
                 """
                cursor.execute(query, (
                Session.last_message_id, Session.current_user["id_user"], Session.loaded_chat, Session.loaded_chat,
                Session.current_user["id_user"]))
                results = cursor.fetchall()
                if results:
                    query = """
                         UPDATE mensagens_chat
                         SET lida = 1
                         WHERE lida = 0 AND destinatario_id = %s AND remetente_id = %s
                     """
                    cursor.execute(query, (Session.current_user["id_user"], Session.loaded_chat))
                    db.commit()
                    print(query)

                    self.new_data.emit(results)

                db.close()
                cursor.close()
                time.sleep(2)

        def stop(self):
            self.running = False

    def __init__(self, stacked_widget):
        super().__init__()
        loadUi(resource_path("design/NOVODASH.ui"), self)
        self.widget = stacked_widget
        self.alterar = None
        self.licaoAlterar = None
        self.dados = None

        Session.last_message_id = 0

        self.stack = self.findChild(QWidget, "stackedwidget_btns_da_sidebar")
        self.dashboardStack = self.findChild(QWidget, "stacked_widget_botoes_principais_do_dashboard")
        self.btn_sair.clicked.connect(self.logOut)

        self.stackList = [
            self.findChild(QPushButton, "btn_home"),
            self.findChild(QPushButton, "btn_chat"),
            self.findChild(QPushButton, "btn_cadastro"),
            self.findChild(QPushButton, "btn_perfil"),
        ]

        self.dashboardList = [
            self.findChild(QPushButton, "btn_rendimento_do_dash"),
            self.findChild(QPushButton, "btn_licoes_do_dash"),
            self.findChild(QPushButton, "btn_func_do_dash"),
            self.findChild(QPushButton, "btn_calendario_do_dash"),
        ]

        from functools import partial
        for i, botao in enumerate(self.stackList):
            botao.clicked.connect(partial(self.mudarTela, i))

        for i, botao in enumerate(self.dashboardList):
            botao.clicked.connect(partial(self.mudarDashboard, i))

        self.ListUsers()

        self.chat_timer = self.DBLoopUpdate()
        self.chat_timer.new_data.connect(self.updateChat)
        self.chat_timer.start()

        self.btn_confirmar.clicked.connect(self.cadastrarUsuario)
        self.btn_enviar.clicked.connect(self.sendMessage)
        self.btn_selecionar_foto.clicked.connect(self.escolherFoto)

        self.foto_func.setPixmap(Session.current_user["foto_perfil"])
        self.nome_func.setText(Session.current_user["nome"])
        self.carg_func.setText(Session.current_user["cargo"])

        self.updateUserTable()
        self.atualizarLicoes()
        self.atualizarPerfil(Session.current_user["id_user"])

        self.btn_voltar.clicked.connect(partial(self.mudarDashboard, 1))
        self.btn_concluir.clicked.connect(self.cadastrarLicao)
        self.btn_visualizar_perfil.clicked.connect(lambda: self.atualizarPerfil(Session.loaded_chat))
        self.btn_editar_perfil.clicked.connect(self.abrirSobreMim)
        self.btn_cancelar.clicked.connect(self.cancelarCadastro)

        self.btn_adicionar_cargo.clicked.connect(self.abrirCargo)
        self.btn_excluir_cargo.clicked.connect(self.excluirCargo)

        self.btn_adicionar_depto.clicked.connect(self.abrirDepto)
        self.btn_excluir_depto.clicked.connect(self.excluirDepto)

        self.btn_anexar.clicked.connect(self.anexarLicao)

        self.widget_calendario.activated.connect(self.calendarioClique)
        self.btn_baixar_tabela.clicked.connect(self.exportar_para_excel)

        self.proximo.clicked.connect(self.proximaPagina)
        self.voltar.clicked.connect(self.paginaAnterior)

        self.atualizarCargoDepto()
        self.atualizarCalendario()
        self.atualizarDashboard()
        self.atualizarGrafico()


    def exportar_para_excel(self):
        # Se sua tabela for QTableWidget:
        table = self.tabela_usuarios

        # Abrir janela para escolher onde salvar
        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar como Excel",
            "",
            "Arquivo Excel (*.xlsx)"
        )

        if not caminho:
            return

        wb = Workbook()
        ws = wb.active

        # ------ Cabeçalhos ------
        col_count = table.columnCount()
        row_count = table.rowCount()

        for col in range(col_count):
            header = table.horizontalHeaderItem(col)
            ws.cell(row=1, column=col + 1).value = header.text() if header else ""

        # ------ Dados ------
        for row in range(row_count):
            for col in range(col_count):
                item = table.item(row, col)
                ws.cell(row=row + 2, column=col + 1).value = item.text() if item else ""

        # Salvar arquivo
        try:
            wb.save(caminho)
            QMessageBox.information(self, "Sucesso", "Arquivo Excel salvo com sucesso!")
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao salvar arquivo:\n{e}")

    def updateDashboardList(self):
        db = bancoDados.conectar()
        if not db:
            return []

        cursor = db.cursor()

        # pegar total de licoes
        cursor.execute("SELECT COUNT(*) FROM licoes")
        total_licoes = cursor.fetchone()[0]

        # menos de 75%
        query_menos = """
            SELECT 
                u.id_user,
                u.nome,
                u.foto_perfil,
                COUNT(ul.id_licao) AS concluidas,
                ROUND((COUNT(ul.id_licao) / %s) * 100, 2) AS porcentagem
            FROM usuario u
            LEFT JOIN usuario_licao_realizada ul
                   ON u.id_user = ul.id_usuario
            WHERE u.tipo_usuario = 'user'
            GROUP BY u.id_user
            HAVING porcentagem < 60
            ORDER BY porcentagem DESC
        """
        cursor.execute(query_menos, (total_licoes,))
        usuarios_menos_75 = cursor.fetchall()

        # mais de 75%
        query_mais = """
            SELECT 
                u.id_user,
                u.nome,
                u.foto_perfil,
                COUNT(ul.id_licao) AS concluidas,
                ROUND((COUNT(ul.id_licao) / %s) * 100, 2) AS porcentagem
            FROM usuario u
            LEFT JOIN usuario_licao_realizada ul
                   ON u.id_user = ul.id_usuario
            WHERE u.tipo_usuario = 'user'
            GROUP BY u.id_user
            HAVING porcentagem >= 60
            ORDER BY porcentagem DESC
        """
        cursor.execute(query_mais, (total_licoes,))
        usuarios_mais_75 = cursor.fetchall()

        cursor.close()
        db.close()

        usersAcima = []
        usersAbaixo = []

        for id, nome, foto, concluidas, porcentagem in usuarios_mais_75:
            pixmap = QPixmap()

            if foto:
                pixmap.loadFromData(foto)

            if pixmap.isNull():
                pixmap = QPixmap(resource_path("imagens/user.png"))

            usersAcima.append({
                "id_user": id,
                "nome": nome,
                "foto_perfil": pixmap,
                "concluidas" : concluidas,
                "porcentagem" : porcentagem,
            })

        for id, nome, foto, concluidas, porcentagem in usuarios_menos_75:
            pixmap = QPixmap()

            if foto:
                pixmap.loadFromData(foto)

            if pixmap.isNull():
                pixmap = QPixmap(resource_path("imagens/user.png"))

            usersAbaixo.append({
                "id_user": id,
                "nome": nome,
                "foto_perfil": pixmap,
                "concluidas" : concluidas,
                "porcentagem" : porcentagem,
            })

        return usersAcima, usersAbaixo

    def atualizarGrafico(self):
        db = bancoDados.conectar()
        cursor = db.cursor()

        query = """
            SELECT DATE_FORMAT(data_registro, '%m/%Y') AS mes, COUNT(*) AS total
            FROM usuario_licao_realizada
            GROUP BY mes
            ORDER BY mes;
        """
        cursor.execute(query)
        dados = cursor.fetchall()

        cursor.close()
        db.close()

        self.meses = [linha[0] for linha in dados]
        self.totais = [linha[1] for linha in dados]

        self.items_por_pagina = 6
        self.pagina_atual = 0

        self.desenharGrafico()

    def proximaPagina(self):
        max_paginas = len(self.meses) // self.items_por_pagina
        if self.pagina_atual < max_paginas:
            self.pagina_atual += 1
            self.desenharGrafico()

    def paginaAnterior(self):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self.desenharGrafico()

    def desenharGrafico(self):
        inicio = self.pagina_atual * self.items_por_pagina
        fim = inicio + self.items_por_pagina

        meses_pagina = self.meses[inicio:fim]
        totais_pagina = self.totais[inicio:fim]

        # Remover gráfico antigo
        layout = self.grafico_dash.layout()
        if layout is None:
            layout = QVBoxLayout(self.grafico_dash)
            self.grafico_dash.setLayout(layout)

        for i in reversed(range(layout.count())):
            widget = layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        fig = Figure(figsize=(6, 3), facecolor="none")
        ax = fig.add_subplot(111, facecolor="none")

        # ====================
        #     CORES DAS BARRAS
        # ====================

        cores = plt.cm.tab20(range(len(totais_pagina)))

        ax.bar(meses_pagina, totais_pagina, width=0.15, color=cores)

        # Impedir esticar horizontal
        ax.set_xlim(-1, len(meses_pagina))

        # Ajustar eixo Y sempre 1 acima do máximo
        max_val = max(totais_pagina) if totais_pagina else 0
        ax.set_ylim(0, max_val + 1)

        # ====================
        #       ESTILO
        # ====================

        # Fundo invisível
        fig.patch.set_alpha(0)
        ax.patch.set_alpha(0)

        # Texto branco
        ax.title.set_color("white")
        ax.xaxis.label.set_color("white")
        ax.yaxis.label.set_color("white")

        # Texto branco e NEGRITO
        ax.set_xlabel("Meses", color="white", fontweight="bold")
        ax.set_ylabel("Lições", color="white", fontweight="bold")
        ax.set_title("Progresso Mensal", color="white", fontweight="bold")

        # Eixos em branco e negrito
        ax.tick_params(axis='x', colors='white', labelsize=9)
        ax.tick_params(axis='y', colors='white', labelsize=9)

        # Bordas brancas
        for spine in ax.spines.values():
            spine.set_color("white")

        fig.tight_layout()

        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

    def atualizarDashboard(self):
        container = self.scrollArea_inativos.widget()
        self.scrollArea_inativos.setWidgetResizable(True)

        layout = container.layout()
        self.clearLayout(layout)

        usersAcima, usersAbaixo = self.updateDashboardList()

        for user in usersAbaixo:
            btn = usuarioChat(user)

            btn.mensagens.setVisible(False)

            icon = user["foto_perfil"].scaled(btn.foto_cntt.size(), Qt.KeepAspectRatioByExpanding,
                                              Qt.SmoothTransformation)
            btn.foto_cntt.setFixedSize(60, 60)
            btn.foto_cntt.setPixmap(icon)

            path = QPainterPath()
            path.addEllipse(QRectF(0, 0, 60, 60))
            region = QRegion(path.toFillPolygon().toPolygon())

            btn.foto_cntt.setMask(region)
            btn.online.setText(f'Concluídas: {user["concluidas"]} ({user["porcentagem"]}%)')

            layout.addWidget(btn)

        layout.addStretch()

        container = self.scrollArea_ativos.widget()
        self.scrollArea_ativos.setWidgetResizable(True)

        layout = container.layout()
        self.clearLayout(layout)

        for user in usersAcima:
            btn = usuarioChat(user)

            btn.mensagens.setVisible(False)

            icon = user["foto_perfil"].scaled(btn.foto_cntt.size(), Qt.KeepAspectRatioByExpanding,
                                              Qt.SmoothTransformation)
            btn.foto_cntt.setFixedSize(60, 60)
            btn.foto_cntt.setPixmap(icon)

            path = QPainterPath()
            path.addEllipse(QRectF(0, 0, 60, 60))
            region = QRegion(path.toFillPolygon().toPolygon())

            btn.foto_cntt.setMask(region)
            btn.online.setText(f'Concluídas: {user["concluidas"]} ({user["porcentagem"]}%)')

            layout.addWidget(btn)

        layout.addStretch()

    def anexarLicao(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Selecione o arquivo", "", "Arquivos (*.pdf *.mp4 *.mkv *.avi)"
        )

        if not file_path:
            return

        tamanho = os.path.getsize(file_path) / (1024 * 1024)
        if tamanho > 200:
            QMessageBox.warning(self, "Arquivo grande demais",
                                f"O arquivo tem {tamanho:.2f} MB.\nO limite é 200 MB.")
            return

        nome = os.path.basename(file_path)
        tipo = os.path.splitext(nome)[1].lower().replace('.', '')
        self.label_anexo.setText(nome)

        with open(file_path, "rb") as f:
            self.dados = [f.read(), tipo, nome]
            QMessageBox.information(self, "Sucesso", f"Arquivo '{nome}' enviado com sucesso!")

    def atualizarCalendario(self):
        db = bancoDados.conectar()

        if not db:
            return

        try:
            cursor = db.cursor()
            cursor.execute("SELECT data FROM calendario")
            datas = cursor.fetchall()
            cursor.close()
            db.close()

            formato_verde = QTextCharFormat()
            formato_verde.setBackground(QBrush(QColor("#90EE90")))  # Verde claro
            formato_verde.setForeground(QBrush(QColor("#000000")))

            self.widget_calendario.setDateTextFormat(QDate(), QTextCharFormat())

            # Marca as datas salvas
            for (data_mysql,) in datas:
                data_qt = QDate(data_mysql.year, data_mysql.month, data_mysql.day)
                self.widget_calendario.setDateTextFormat(data_qt, formato_verde)

        except mc.Error as e:
            QMessageBox.critical(self, "Erro", f"Erro ao buscar datas: {e}")

    def calendarioClique(self, date: QDate):
        data_str = date.toString("yyyy-MM-dd")
        db = bancoDados.conectar()
        if not db:
            return

        cursor = db.cursor()
        cursor.execute("SELECT * FROM calendario WHERE data = %s", (data_str,))
        resultado = cursor.fetchone()

        if resultado is None:
            popup = PopupCalendario(self)
            popup.label_data.setText(date.toString("dd-MM-yyyy"))
            resultado_popup = popup.exec_()

            if resultado_popup == QDialog.Accepted:
                valores = popup.valor_retornado

                if valores[0] == "" or valores[1] == "":
                    QMessageBox.warning(self, "Aviso", "Preencha todos os campos!")
                    return

                query_insert = """
                    INSERT INTO calendario (titulo, descricao, data)
                    VALUES (%s, %s, %s)
                """
                try:
                    cursor.execute(query_insert, (valores[0], valores[1], data_str))
                    db.commit()
                    QMessageBox.information(self, "Sucesso", "Adicionado com sucesso!")
                    self.atualizarCalendario()
                except mc.Error as err:
                    QMessageBox.critical(self, "Erro", f"Erro ao inserir no banco: {err}")
            else:
                print("Usuário cancelou o popup.")

        else:
            id_evento, titulo, descricao, data = resultado

            msg = QMessageBox()
            msg.setWindowTitle("Opção de Calendário")
            msg.setText(
                f"Já existe um compromisso nesta data:\n\n📅 {titulo}\n\nVocê gostaria de visualizar ou excluir o compromisso?")
            msg.setIcon(QMessageBox.Question)

            btn_visualizar = msg.addButton("Visualizar", QMessageBox.ActionRole)
            btn_excluir = msg.addButton("Excluir", QMessageBox.DestructiveRole)
            btn_cancelar = msg.addButton("Cancelar", QMessageBox.RejectRole)
            msg.setDefaultButton(btn_visualizar)

            msg.exec_()

            if msg.clickedButton() == btn_visualizar:
                print(f"Usuário escolheu visualizar o compromisso {id_evento}")
                popup = PopupVisualizarCal(self)
                popup.label_data.setText(date.toString("dd-MM-yyyy"))
                popup.lineEdit_titulo.setText(titulo)
                popup.textEdit_descricao.setPlainText(descricao)
                resultado_popup = popup.exec_()

            elif msg.clickedButton() == btn_excluir:
                print(f"Usuário escolheu excluir o compromisso {id_evento}")
                confirmar = QMessageBox.question(
                    self,
                    "Confirmação",
                    "Tem certeza que deseja excluir este compromisso?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if confirmar == QMessageBox.Yes:
                    cursor.execute("DELETE FROM calendario WHERE id_calendario = %s", (id_evento,))
                    db.commit()
                    QMessageBox.information(self, "Sucesso", "Compromisso excluído com sucesso!")
                    self.atualizarCalendario()
            else:
                print("Usuário cancelou a ação.")

        cursor.close()
        db.close()

    def on_alterar(self, user_id):
        print(f"Alterar usuário {user_id}")
        self.alterar = user_id
        self.btn_cadastro.setChecked(True)
        self.mudarTela(2)

    def on_excluir(self, user_id, user_name):
        reply = QMessageBox.question(
            self,
            "Confirmação",
            f"Tem certeza que deseja excluir o usuário {user_name} (ID {user_id})?",
            QMessageBox.No | QMessageBox.Yes,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            print(f"Usuário {user_id} excluído")

            ifdb = bancoDados.conectar()
            ifCursor = ifdb.cursor()

            query = """
                        DELETE FROM usuario WHERE id_user = %s;
                    """
            ifCursor.execute(query, (user_id,))
            ifdb.commit()
            ifCursor.close()
            ifdb.close()

            QMessageBox.information(self, "Sucesso", f"Usuário {user_name} excluido com sucesso!")
            self.updateUserTable()
        else:
            QMessageBox.information(self, "Cancelado", f"Exclusão cancelada.")
            print(f"Exclusão do usuário {user_id} cancelada")

    def pixmap_to_bytes(self, pixmap):
        if pixmap is None:
            return None
        ba = QByteArray()
        buffer = QBuffer(ba)
        buffer.open(QIODevice.WriteOnly)
        pixmap.save(buffer, "PNG")
        return ba.data()

    def abrirDepto(self):

        popup = PopupDepto(self)
        resultado = popup.exec_()  # Abre o popup de forma modal

        # Se o usuário confirmou (clicou em "Confirmar")
        if resultado == QDialog.Accepted:
            print("eba")
            valores = popup.valor_retornado
            db = bancoDados.conectar()

            if valores[0] == "" or valores[1] == "":
                QMessageBox.warning(self, "Aviso", "Preencha todos os campos!")
                return

            cursor = db.cursor()

            query = "SELECT * FROM departamento WHERE nome_depto = %s"
            cursor.execute(query, (valores[0],))
            result = cursor.fetchone()

            print("AAAAAAAAAAAAAAAAAAAAAAAA")

            if result is None:

                query = """
                    INSERT INTO departamento (nome_depto, foto_depto)
                    VALUES (%s, %s)
                """
                try:
                    cursor = db.cursor()
                    cursor.execute(query, (valores[0], self.pixmap_to_bytes(valores[1])))
                    db.commit()
                    cursor.close()
                    db.close()

                    QMessageBox.information(self, "Sucesso", "Adicionado com sucesso!")
                    self.atualizarCargoDepto()
                except mc.Error as err:
                    print("Error:", err)

            else:
                QMessageBox.warning(None, "Aviso", "Cargo já existente.")

        else:
            print("Usuário cancelou o popup.")

    def abrirCargo(self):
        popup = PopupCargo(self)
        resultado = popup.exec_()  # Abre o popup de forma modal

        # Se o usuário confirmou (clicou em "Confirmar")
        if resultado == QDialog.Accepted:
            print("eba")
            valores = popup.valor_retornado
            db = bancoDados.conectar()

            if valores[0] == "" or valores[1] == "":
                QMessageBox.warning(self, "Aviso", "Preencha todos os campos!")
                return

            cursor = db.cursor()

            query = f"SELECT * FROM cargo WHERE nome_cargo = %s"
            cursor.execute(query, (valores[0],))
            result = cursor.fetchone()

            if result is None:
                values = {"Usuário": "user", "Administrador": "admin"}

                query = """
                                                     INSERT INTO cargo
                                                     (nome_cargo, permissao_cargo) 
                                                     VALUES (%s, %s);
                                                   """
                try:
                    cursor = db.cursor()
                    cursor.execute(query, (valores[0], values[valores[1]]))
                    db.commit()
                    cursor.close()
                    db.close()

                    QMessageBox.information(self, "Sucesso", "Adicionado com sucesso!")
                    self.atualizarCargoDepto()
                except mc.Error as err:
                    print("Error:", err)

            else:
                QMessageBox.warning(None, "Aviso", "Cargo já existente.")

        else:
            print("Usuário cancelou o popup.")

    def abrirSobreMim(self):
        popup = PopupSobreMim(self)
        resultado = popup.exec_()  # Abre o popup de forma modal

        # Se o usuário confirmou (clicou em "Confirmar")
        if resultado == QDialog.Accepted:
            print("eba")
            valores = popup.valor_retornado
            db = bancoDados.conectar()

            if valores[0] == "" or valores[1] == "":
                QMessageBox.warning(self, "Aviso", "Preencha todos os campos!")
                return

            query = """
                           UPDATE usuario SET sobre_mim = %s, experiencias = %s WHERE id_user = %s;
                      """
            try:
                cursor = db.cursor()
                cursor.execute(query, (valores[0], valores[1], Session.current_user["id_user"]))
                db.commit()
                cursor.close()
                db.close()

                QMessageBox.information(self, "Sucesso", "Perfil atualizado!")
                self.atualizarPerfil(Session.current_user["id_user"])

            except mc.Error as err:
                print("Error:", err)
        else:
            print("Usuário cancelou o popup.")

    def atualizarCargoDepto(self):
        db = bancoDados.conectar()
        cursor = db.cursor()
        cursor.execute("SELECT * FROM cargo")

        rows = cursor.fetchall()
        self.comboBox_cargo.clear()

        for id_cargo, nome_cargo, permissao_cargo in rows:
            self.comboBox_cargo.addItem(nome_cargo)

        cursor.execute("SELECT * FROM departamento")

        rows = cursor.fetchall()
        self.comboBox_depto.clear()

        for id_depto, nome_depto, foto_depto in rows:
            self.comboBox_depto.addItem(nome_depto)

    def excluirCargo(self):
        if self.comboBox_cargo.currentText() is None or self.comboBox_cargo.currentText() == "":
            QMessageBox.warning(self, "Aviso", f"Cargo inexistente.")
            return

        reply = QMessageBox.question(
            self,
            "Confirmação",
            f"Tem certeza que deseja excluir este cargo?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            ifdb = bancoDados.conectar()
            ifCursor = ifdb.cursor()

            query = """
                               DELETE FROM cargo WHERE nome_cargo = %s;
                           """
            ifCursor.execute(query, (self.comboBox_cargo.currentText(), ))
            ifdb.commit()
            ifCursor.close()
            ifdb.close()

            QMessageBox.information(self, "Sucesso", f"Cargo {self.comboBox_cargo.currentText()} deletado com sucesso!")

            self.atualizarCargoDepto()

    def excluirDepto(self):
        if self.comboBox_depto.currentText() is None or self.comboBox_depto.currentText() == "":
            QMessageBox.warning(self, "Aviso", f"Departamento inexistente.")
            return

        reply = QMessageBox.question(
            self,
            "Confirmação",
            f"Tem certeza que deseja excluir este departamento?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            ifdb = bancoDados.conectar()
            ifCursor = ifdb.cursor()

            query = """
                               DELETE FROM departamento WHERE nome_depto = %s;
                           """
            ifCursor.execute(query, (self.comboBox_depto.currentText(),))
            ifdb.commit()
            ifCursor.close()
            ifdb.close()

            QMessageBox.information(self, "Sucesso", f"Departamento {self.comboBox_depto.currentText()} deletado com sucesso!")

            self.atualizarCargoDepto()

    def alterarLicao(self, user_id):
        print(f"Alterar licao {user_id}")
        self.licaoAlterar = user_id
        self.label_anexo.setText("PDF;VIDEO_AULA")
        self.dados = None
        self.mudarDashboard(4)

    def excluirLicao(self, user_id, user_name):
        reply = QMessageBox.question(
            self,
            "Confirmação",
            f"Tem certeza que deseja excluir a lição {user_name} (ID {user_id})?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            print(f"Usuário {user_id} excluído")

            ifdb = bancoDados.conectar()
            ifCursor = ifdb.cursor()

            query = """
                               DELETE FROM licoes WHERE id_licao = %s;
                           """
            ifCursor.execute(query, (user_id, ))

            query = """
                               DELETE FROM usuario_licao_realizada WHERE id_licao = %s;
                           """
            ifCursor.execute(query, (user_id, ))
            ifdb.commit()
            ifCursor.close()
            ifdb.close()

            QMessageBox.information(self, "Sucesso", f"Lição {user_name} excluida com sucesso!")
            self.atualizarLicoes()
        else:
            QMessageBox.information(self, "Cancelado", f"Exclusão cancelada.")
            print(f"Exclusão da lição {user_id} cancelada")

    def atualizarPerfil(self, id):
        print(id)
        db = bancoDados.conectar()
        cursor = db.cursor()
        cursor.execute("SELECT id_user, nome, departamento, cargo, foto_perfil, sobre_mim, experiencias FROM usuario WHERE id_user = %s", (id, ))

        row = cursor.fetchone()

        if not (row is None):
            id_user, nome, departamento, cargo, foto_perfil, sobre_mim, experiencias = row

            self.btn_editar_perfil.setVisible(True if Session.current_user["id_user"] == id_user else False)

            self.nome_funcionario.setText(f'<html><head/><body><p><span style=" font-size:22pt;">{nome}</span></p></body></html>')
            self.cargo_func.setText(f'<html><head/><body><p><span style=" font-size:14pt;">{cargo}</span></p></body></html>')
            self.sobre_mim.setText(f'<html><head/><body><p><span style=" font-size:10pt;">{sobre_mim}</span></p></body></html>')
            self.experiencias.setText(f'<html><head/><body><p><span style=" font-size:9pt;">{experiencias}</span></p></body></html>')

            pixmap = QPixmap()
            pixmap.loadFromData(foto_perfil)  # converte bytes → QPixmap

            if pixmap.isNull():
                pixmap = QPixmap(resource_path("imagens/user.png"))

            self.foto_funcionario.setPixmap(pixmap)

            if Session.loaded_chat:
                self.mudarTela(3, False)
                self.btn_perfil.setChecked(True)

            cursor.execute("SELECT foto_depto FROM departamento WHERE nome_depto = %s", (departamento,))
            row = cursor.fetchone()
            if row and row[0]:
                image_data = row[0]
                print(type(image_data), len(image_data))
                print(image_data[:20])
                pixmap = QPixmap()
                pixmap.loadFromData(image_data)  # converte bytes → QPixmap

                if pixmap.isNull():
                    pixmap = QPixmap(resource_path("imagens/logo new te.png"))

                self.foto_aqui.setPixmap(pixmap)  # coloca na QLabel
                self.foto_aqui.setScaledContents(True)  # ajusta o tamanho automaticamente
                print("Imagem carregada com sucesso!")
            else:
                print("Nenhuma imagem encontrada para esse departamento.")
                pixmap = QPixmap(resource_path("imagens/logo new te.png"))

                self.foto_aqui.setPixmap(pixmap)  # coloca na QLabel
                self.foto_aqui.setScaledContents(True)  # ajusta o tamanho automaticamente

    def atualizarLicoes(self):
        container = self.scroll_licoes.widget()

        self.scroll_licoes.setWidgetResizable(True)
        layout = container.layout()

        self.clearLayout(layout)

        db = bancoDados.conectar()
        cursor = db.cursor()
        cursor.execute("SELECT * FROM licoes")

        rows = cursor.fetchall()

        i = 0

        for id_licao, id_user, titulo, desc, metas, criacao, validade, nomeArquivo, arquivo, tipo in rows:
            row = i // 4
            col = i % 4
            template = licao()
            layout.addWidget(template, row, col)
            template.lbl_titulo_curso.setText(titulo)
            template.lbl_desc_curso.setText(desc)
            template.lbl_tipo_de_arquivo.setText("Nenhum" if (not nomeArquivo or not arquivo) else f"{nomeArquivo}")
            template.btn_excluir.clicked.connect(partial(self.excluirLicao,id_licao, titulo))
            template.btn_editar.clicked.connect(partial(self.alterarLicao, id_licao))
            i = i + 1

        row = i // 4
        col = i % 4

        self.adicionar_licao = adicionarLicao()
        layout.addWidget(self.adicionar_licao, row, col)

        self.adicionar_licao.botao.clicked.connect(partial(self.mudarDashboard, 4, True))

    def updateUserTable(self):
        self.tabela_usuarios.clear()

        db = bancoDados.conectar()
        cursor = db.cursor()
        cursor.execute("SELECT * FROM usuario WHERE tipo_usuario = 'user'")

        rows = cursor.fetchall()
        col_names = [desc[0] for desc in cursor.description]

        excluded = ["foto_perfil", "experiencias", "sobre_mim", "senha"]
        exclude_indices = [i for i, name in enumerate(col_names) if name in excluded]

        visible_col_names = [name for i, name in enumerate(col_names) if i not in exclude_indices]

        self.tabela_usuarios.setRowCount(len(rows))
        self.tabela_usuarios.setColumnCount(len(visible_col_names) + 2)
        self.tabela_usuarios.setHorizontalHeaderLabels(visible_col_names + [" ", "  "])

        self.alterar = None

        for r, row in enumerate(rows):
            col_position = 0

            for c, value in enumerate(row):
                if c in exclude_indices:
                    continue
                self.tabela_usuarios.setItem(r, col_position, QTableWidgetItem(str(value)))
                col_position += 1

            user_id = row[0]
            user_name = row[1]

            # Botão Alterar
            btn_alterar = QPushButton("Alterar")
            btn_alterar.setStyleSheet("background-color: yellow; font-weight: bold; color: black;")
            btn_alterar.clicked.connect(lambda _, uid=user_id: self.on_alterar(uid))

            btn_alterar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.tabela_usuarios.setCellWidget(r, col_position, btn_alterar)
            col_position += 1

            btn_excluir = QPushButton("Excluir")
            btn_excluir.setStyleSheet("background-color: red; font-weight: bold; color: white;")
            btn_excluir.clicked.connect(lambda _, uid=user_id, uname=user_name: self.on_excluir(uid, uname))

            btn_excluir.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.tabela_usuarios.setCellWidget(r, col_position, btn_excluir)

            self.tabela_usuarios.resizeRowToContents(r)

            self.tabela_usuarios.setColumnWidth(self.tabela_usuarios.columnCount() - 2, 90)
            self.tabela_usuarios.setColumnWidth(self.tabela_usuarios.columnCount() - 1, 90)

        cursor.close()
        db.close()

    def ListUsers(self):
        container = self.usuarios_chat.widget()
        self.usuarios_chat.setWidgetResizable(True)

        layout = container.layout()
        self.clearLayout(layout)

        users = self.updateUserList()

        def callback(user):
            Session.loaded_chat = user["id_user"]

            path = QPainterPath()
            path.addEllipse(QRectF(0, 0, 60, 60))
            region = QRegion(path.toFillPolygon().toPolygon())

            self.infos_contato.setText(user["nome"])

            #self.foto_func_contato.setPixmap(user["foto_perfil"])
            #self.foto_func_contato.setMask(region)

            #self.foto_func_contato_2.setPixmap(user["foto_perfil"])
            #self.foto_func_contato_2.setMask(region)

            self.nome_func_contato.setText(user["nome"])
            self.carg_func_contato.setText(user["cargo"])

            Session.last_message_id = 0

            self.clearLayout(self.chat.widget().layout())
            self.chat_timer.query()

        for user in users:
            btn = usuarioChat(user)
            btn.pushButton.clicked.connect(partial(callback, user))

            btn.mensagens.setVisible(False)

            if user["mensagens"] > 0:
                btn.mensagens.setVisible(True)
                btn.mensagens.setText(f'{user["mensagens"]}')

            icon = user["foto_perfil"].scaled(btn.foto_cntt.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            btn.foto_cntt.setFixedSize(60, 60)
            btn.foto_cntt.setPixmap(icon)

            path = QPainterPath()
            path.addEllipse(QRectF(0, 0, 60, 60))
            region = QRegion(path.toFillPolygon().toPolygon())

            btn.foto_cntt.setMask(region)

            btn.online.setText(user["status"])

            if user["status"] == "OFFLINE":
                btn.online.setStyleSheet("""
                    QLabel {
                        color: red;
                    }
                """)
            else:
                btn.online.setStyleSheet("""
                    QLabel {
                        color: green;
                    }
                """)

            layout.addWidget(btn)

        layout.addStretch()


    def escolherFoto(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Image", "", "Images (*.png *.jpg *.jpeg *.bmp *.jfif)"
        )

        if not file_path:
            return  # user cancelled

        pixmap = QPixmap(file_path)
        if pixmap.isNull():
            QMessageBox.warning(self, "Erro", "Não foi possível carregar a imagem!")
            return

        # Set button icon
        icon = QIcon(pixmap)
        self.foto_novo_func.setIcon(icon)

    def updateUserList(self):
        db = bancoDados.conectar()
        if not db:
            return []

        cursor = db.cursor()

        query = """
              SELECT
                  u.id_user,
                  u.nome,
                  u.foto_perfil,
                  u.cargo,
                  u.status,
                  (
                      SELECT MAX(m.data_envio)
                      FROM mensagens_chat m
                      WHERE 
                          (m.remetente_id = u.id_user AND m.destinatario_id = %s)
                          OR
                          (m.destinatario_id = u.id_user AND m.remetente_id = %s)
                  ) AS data_envio
              FROM usuario u
              WHERE u.id_user != %s
              ORDER BY data_envio DESC
          """
        current_id = Session.current_user["id_user"]

        cursor.execute(query, (current_id, current_id, current_id))
        results = cursor.fetchall()

        users = []
        for id_user, nome, foto_perfil, cargo, status, ultima_msg in results:
            pixmap = QPixmap()

            if foto_perfil:
                pixmap.loadFromData(foto_perfil)

            if pixmap.isNull():
                pixmap = QPixmap(resource_path("imagens/user.png"))

            query = "SELECT COUNT(*) FROM mensagens_chat WHERE destinatario_id = %s AND lida = 0 AND remetente_id = %s"
            cursor.execute(query, (Session.current_user["id_user"], id_user))
            mensagens = cursor.fetchone()[0]

            users.append({
                "id_user": id_user,
                "nome": nome,
                "foto_perfil": pixmap,
                "cargo" : cargo,
                "status" : status,
                "mensagens" : mensagens,
            })
            print(ultima_msg)

        cursor.close()
        db.close()

        return users

    def clearLayout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clearLayout(item.layout())

    def sendMessage(self):
        text =  filtrar_texto(self.lineEdit_mensagem.text())
        db = bancoDados.conectar()

        if text == "" or text is None:
            return

        self.lineEdit_mensagem.setText("")

        query = """
          INSERT INTO mensagens_chat
          (remetente_id, destinatario_id, mensagem) 
          VALUES (%s, %s, %s);
          """
        try:
            cursor = db.cursor()
            cursor.execute(query, (Session.current_user["id_user"], Session.loaded_chat, text))
            db.commit()
            cursor.close()
            db.close()

            self.chat_timer.query()
        except mc.Error as err:
            print("Error:", err)

    def updateChat(self, results):
        self.ListUsers()
        if results:
            Session.last_message_id = results[len(results)-1][0]

            self.chat.setWidgetResizable(True)
            container = self.chat.widget()
            layout = container.layout()
            layout.setAlignment(Qt.AlignTop)

            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(5)

            if layout.count() == 0:
                layout.addStretch()

            container.setMinimumHeight(self.chat.viewport().height())

            for row in results:
                bubble = ChatBubble(
                    row[3],
                    self.chat.viewport().width() / 2,
                    sender="me" if row[1] == Session.current_user["id_user"] else "other",
                )
                layout.insertWidget(layout.count() - 1, bubble)
                bubble.adjustSize()

            #container.setMinimumHeight(self.chat.viewport().height())

            self.resize(self.width() + 1, self.height())
            self.chat.setMaximumHeight(920)
            self.scrollToBottom()

    def scrollToBottom(self):
        # Delay slightly to ensure layout has fully recalculated
        QTimer.singleShot(50, self._scrollToBottomImmediate)

    def _scrollToBottomImmediate(self):
        sb = self.chat.verticalScrollBar()
        sb.setValue(sb.maximum())
        self.resize(self.width() - 1, self.height())

    def cancelarCadastro(self):
        self.limparCampos()
        self.alterar = None
        self.mudarTela(2)

    def mudarTela(self, index, update=True):
        if index == 3 and update:
            self.atualizarPerfil(Session.current_user["id_user"])
        if index == 1:
            self.ListUsers()
        self.stack.setCurrentIndex(index)

        if index == 2 and not (self.alterar is None):
            print(f"alterar: {self.alterar}")

            db = bancoDados.conectar()
            if not db:
                return

            cursor = db.cursor()
            query = "SELECT nome, email, telefone, endereco, cpf, rg, departamento, cargo, foto_perfil FROM usuario WHERE id_user = %s"
            cursor.execute(query, (self.alterar,))
            result = cursor.fetchone()

            if not (result is None):
                nome, email, telefone, endereco, cpf, rg, departamento, cargo, foto_perfil = result

                self.titulo_cadastro_2.setText(f"Alterando {nome} ({self.alterar})")

                self.lineEdit_nome.setText(nome)
                self.lineEdit_email.setText(email)
                self.lineEdit_telefone.setText(telefone)
                self.lineEdit_endereco.setText(endereco)
                self.lineEdit_cpf.setText(cpf)
                self.lineEdit_rg.setText(rg)
                self.comboBox_depto.setCurrentText(departamento)
                self.comboBox_cargo.setCurrentText(cargo)

                pixmap = QPixmap()

                if foto_perfil:
                    pixmap.loadFromData(foto_perfil)

                # If pixmap is invalid, use a default avatar
                if pixmap.isNull():
                    pixmap = QPixmap(resource_path("imagens/user.png"))

                icon = QIcon(pixmap)
                self.foto_novo_func.setIcon(icon)

            cursor.close()
            db.close()
        else:
            if index == 2 and self.alterar is None:
                self.titulo_cadastro_2.setText("Cadastrar")

    def mudarDashboard(self, index, cadastrar = False):
        if cadastrar:
            self.licaoAlterar = None
            self.label_anexo.setText("PDF;VIDEO_AULA")
            self.dados = None
            self.limparCampos()

        if index == 1:
            self.atualizarLicoes()
        if index == 4 and not (self.licaoAlterar is None):
            db = bancoDados.conectar()
            if not db:
                return

            cursor = db.cursor()
            query = "SELECT titulo, conteudo, metas, nome_arquivo, arquivo, tipo FROM licoes WHERE id_licao = %s"
            cursor.execute(query, (self.licaoAlterar,))
            result = cursor.fetchone()
            if not (result is None):
                titulo, conteudo, metas, nomeArquivo, arquivo, tipo = result

                # self.titulo_cadastro_2.setText(f"Alterando {titulo} ({self.alterar})")

                self.lineEdit_titulo.setText(titulo)
                self.textEdit_desc.setPlainText(conteudo)
                self.textEdit_metas.setPlainText(metas)
                self.label_anexo.setText("PDF;VIDEO_AULA" if (not nomeArquivo or not arquivo) else f"{nomeArquivo}")
                self.dados = [arquivo, tipo, nomeArquivo]

            cursor.close()
            db.close()
        self.atualizarCalendario()
        self.dashboardStack.setCurrentIndex(index)

    def quitProgram(self):
        if Session.current_user is None:
            return
        db = bancoDados.conectar()
        if not db:
            return

        cursor = db.cursor()

        query = """
                       UPDATE usuario SET status = 'OFFLINE' WHERE email = %s;
                  """
        cursor.execute(query, (Session.current_user["email"],))
        db.commit()

        cursor.close()
        db.close()
        Session.current_user = None

    def logOut(self):
        self.chat_timer.stop()
        self.quitProgram()

        self.close()
        self.widget.setCurrentIndex(0)
        self.widget.show()

    def cadastrarLicao(self):
        db = bancoDados.conectar()
        if not db:
            return

        cursor = db.cursor()

        print(Session.current_user["id_user"])

        valuesDictionaries = {
            "id_user": Session.current_user["id_user"],
            "titulo": self.lineEdit_titulo.text(),
            "conteudo": self.textEdit_desc.toPlainText(),
            "metas": self.textEdit_metas.toPlainText(),
            "nome_arquivo": self.dados[2],
            "arquivo": self.dados[0],
            "tipo": self.dados[1]
        }

        print("eba")

        for key, value in valuesDictionaries.items():
            if value is None or value == "":
                QMessageBox.warning(self, "Erro", f"Preencha todos os campos")
                return

        query = ""

        if self.licaoAlterar:
            query = f"""
                  UPDATE licoes
                  SET 
                      id_user = %s,
                      titulo = %s,
                      conteudo = %s,
                      metas = %s,
                      nome_arquivo = %s,
                      arquivo = %s,
                      tipo = %s
                  WHERE id_licao = {self.licaoAlterar};
                  """
        else:
            query = """
                  INSERT INTO licoes
                  (id_user, titulo, conteudo, metas, nome_arquivo, arquivo, tipo)
                  VALUES (%s, %s, %s, %s, %s, %s, %s);
                  """

        values = tuple(valuesDictionaries.values())

        QMessageBox.information(self, "Carregando...",
                                "Enviando tarefa... caso seu anexo seja muito grande, por favor aguarde.")

        try:
            cursor.execute(query, values)
            db.commit()

            QMessageBox.information(self, "Sucesso", "Tarefa alterada com sucesso!" if self.licaoAlterar else "Tarefa cadastrada com sucesso!")

            self.licaoAlterar = None
            self.dados = None
            self.label_anexo.setText("PDF;VIDEO_AULA")
            self.mudarDashboard(1)
            self.limparCampos()
        except mc.Error as err:
            print(f"Falha ao salvar tarefa: {err}")
        finally:
            self.atualizarLicoes()
            cursor.close()
            db.close()

    def limparCampos(self):
        self.lineEdit_nome.clear()
        self.lineEdit_email.clear()
        self.lineEdit_telefone.clear()
        self.lineEdit_cpf.clear()
        self.lineEdit_rg.clear()
        self.comboBox_depto.setCurrentIndex(0)
        self.comboBox_cargo.setCurrentIndex(0)
        self.lineEdit_endereco.clear()

        icon = QIcon(QPixmap(resource_path("imagens/user.png")))
        self.foto_novo_func.setIcon(icon)

        self.lineEdit_titulo.clear()
        self.textEdit_desc.clear()
        self.textEdit_metas.clear()

        self.alterar = None
        self.licaoAlterar = None
        self.dados = None
        self.label_anexo.setText("PDF;VIDEO_AULA")

    def cadastrarUsuario(self):

        db = bancoDados.conectar()
        if not db:
            return

        cursor = db.cursor()

        # Convert pixmap to binary for DB
        pixmap = self.foto_novo_func.icon().pixmap(self.foto_novo_func.iconSize())
        img_data = self.pixmap_to_bytes(pixmap)

        cursor.execute("SELECT permissao_cargo FROM cargo WHERE nome_cargo = %s", (self.comboBox_cargo.currentText(),))
        row = cursor.fetchone()

        cpf, rg = ValidadorCPF(self.lineEdit_cpf.text()).validar(), ValidadorRG(self.lineEdit_rg.text()).validar()

        if not cpf and not rg:
            QMessageBox.warning(self, "Erro", "CPF e RGs inválidos.")
            return
        else:
            if not cpf:
                QMessageBox.warning(self, "Erro", "CPF inválido.")
                return
            else:
                if not rg:
                    QMessageBox.warning(self, "Erro", "RG inválido.")
                    return

        senha = GeradorSenha().gerar()
        senha_bytes = senha.encode("utf-8")
        hash_senha = bcrypt.hashpw(senha_bytes, bcrypt.gensalt())

        receiverEmail =  ValidadorEmail(self.lineEdit_email.text()).validar()
        if not receiverEmail:
            QMessageBox.warning(self, "Erro", "Email inválido.")
            return
        query = f"SELECT email, cpf, rg, senha FROM usuario WHERE email = %s or cpf = %s or rg = %s;"
        cursor.execute(query, (receiverEmail, cpf, rg))
        result = cursor.fetchone()

        if result and not self.alterar:
            QMessageBox.information(self, "Aviso", "Esse E-Mail, CPF ou RG já está cadastrado!")
            return

        valuesDictionaries = {
            "nome": self.lineEdit_nome.text(),
            "email": receiverEmail,
            "telefone": ''.join(c for c in self.lineEdit_telefone.text() if c.isdigit()),
            "cpf": cpf,
            "rg": rg,
            "departamento": self.comboBox_depto.currentText(),
            "cargo": self.comboBox_cargo.currentText(),
            "foto_perfil": img_data,
            "status": "OFFLINE",
            "sobre_mim": "Sobre mim...",
            "senha": result[3] if self.alterar else hash_senha,
            "endereco": self.lineEdit_endereco.text(),
            "tipo_usuario": row[0] if row else "user",
            "experiencias": "Experiências..."
        }

        for key, value in valuesDictionaries.items():
            if value is None or value == "":
                QMessageBox.warning(self, "Erro", f"Preencha todos os campos e foto de perfil!")
                return

        query = ""

        if self.alterar:
            query = f"""
            UPDATE usuario
            SET 
                nome = %s,
                email = %s,
                telefone = %s,
                cpf = %s,
                rg = %s,
                departamento = %s,
                cargo = %s,
                foto_perfil = %s,
                status = %s,
                sobre_mim = %s,
                senha = %s,
                endereco = %s,
                tipo_usuario = %s,
                experiencias = %s
            WHERE id_user = {self.alterar};
            """
        else:
            query = """
            INSERT INTO usuario
            (nome, email, telefone, cpf, rg, departamento, cargo, foto_perfil, status, sobre_mim, senha, endereco, tipo_usuario, experiencias)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """

        values = tuple(valuesDictionaries.values())

        try:
            cursor.execute(query, values)
            db.commit()

            QMessageBox.information(self, "Sucesso", "Usuário alterado com sucesso!" if self.alterar else "Usuário cadastrado com sucesso!" )

            if not self.alterar:
                self.email_thread = EmailSender(receiverEmail, senha)
                self.email_thread.start()

            self.limparCampos()
            self.alterar = None
        except mc.Error as err:
            print(f"Falha ao salvar usuário: {err}")
        finally:
            self.updateUserTable()
            cursor.close()
            db.close()
